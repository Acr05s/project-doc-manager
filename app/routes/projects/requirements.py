"""项目需求配置相关路由"""

import json
import logging
from pathlib import Path
from flask import request, jsonify, Response
from .utils import get_doc_manager
import openpyxl
import tempfile

logger = logging.getLogger(__name__)


def load_project_config():
    """加载项目配置（Excel/JSON文件解析）并保存为独立的requirements文件"""
    try:
        doc_manager = get_doc_manager()
        file = request.files.get('file')
        if not file:
            return jsonify({'status': 'error', 'message': '未获取到文件'}), 400

        # 保存临时文件
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        # 使用 RequirementsLoader 解析文件
        from app.utils.requirements_loader import RequirementsLoader
        loader = RequirementsLoader(doc_manager.config)
        config = loader.load(tmp_path)

        # 删除临时文件
        Path(tmp_path).unlink(missing_ok=True)

        if not config or not config.get('cycles'):
            return jsonify({'status': 'error', 'message': '文件解析失败或格式不正确'}), 400

        # 保存为独立的requirements文件
        logger.info(f"[DEBUG] 调用 save_requirements_config, doc_manager 类型: {type(doc_manager)}")
        logger.info(f"[DEBUG] doc_manager 是否有 save_requirements_config: {hasattr(doc_manager, 'save_requirements_config')}")
        
        result = doc_manager.save_requirements_config(config, file.filename)
        
        # 调试日志
        logger.info(f"[DEBUG] save_requirements_config 返回结果: {result}")
        logger.info(f"[DEBUG] result 类型: {type(result)}")
        if result:
            logger.info(f"[DEBUG] result.keys(): {result.keys() if isinstance(result, dict) else 'N/A'}")
            logger.info(f"[DEBUG] requirements_id: {result.get('requirements_id')}")
            logger.info(f"[DEBUG] name: {result.get('name')}")

        return jsonify({
            'status': 'success',
            'data': config,
            'requirements_id': result.get('requirements_id') if result else None,
            'requirements_name': result.get('name') if result else None
        })
    except Exception as e:
        logger.error(f"[DEBUG] load_project_config 异常: {e}")
        import traceback
        logger.error(f"[DEBUG] 错误堆栈: {traceback.format_exc()}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


def apply_requirements_to_project_route(project_id):
    """将文档需求配置应用到项目"""
    try:
        doc_manager = get_doc_manager()
        data = request.get_json()
        requirements_id = data.get('requirements_id')
        
        logger.info(f"[DEBUG] /apply-requirements 被调用: project_id={project_id}, requirements_id={requirements_id}")

        if not requirements_id:
            return jsonify({'status': 'error', 'message': '未指定需求配置ID'}), 400

        result = doc_manager.apply_requirements_to_project(project_id, requirements_id)
        
        logger.info(f"[DEBUG] apply_requirements_to_project 返回: {result}")
        return jsonify(result)
    except Exception as e:
        logger.error(f"[DEBUG] /apply-requirements 异常: {e}")
        import traceback
        logger.error(f"[DEBUG] 错误堆栈: {traceback.format_exc()}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


def list_requirements_configs():
    """获取所有文档需求配置列表"""
    try:
        doc_manager = get_doc_manager()
        result = doc_manager.list_requirements_configs()
        return jsonify(result)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


def export_requirements():
    """导出需求清单JSON"""
    try:
        doc_manager = get_doc_manager()
        project_id = request.args.get('project_id')
        if not project_id:
            return jsonify({'status': 'error', 'message': '缺少项目ID'}), 400
        
        # 获取项目配置
        project_result = doc_manager.load_project(project_id)
        if project_result['status'] != 'success':
            return jsonify({'status': 'error', 'message': '项目不存在'}), 404
        
        project_config = project_result['project']
        
        # 导出为JSON
        json_content = doc_manager.export_requirements_to_json(project_config)
        
        return Response(
            json_content,
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename=requirements_{project_id}.json'}
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


def get_document_directories(project_id):
    """获取项目的文档目录映射"""
    try:
        from flask_login import current_user
        from app.models.user import user_manager

        cycle = request.args.get('cycle')
        doc_category = request.args.get('doc_category')

        if not cycle or not doc_category:
            return jsonify({'status': 'error', 'message': '缺少周期或文档类型参数'}), 400

        directories = user_manager.get_document_directories(project_id, cycle, doc_category)
        return jsonify({'status': 'success', 'directories': directories})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


def create_document_directory(project_id):
    """创建文档目录映射"""
    try:
        from flask_login import current_user
        from app.models.user import user_manager

        data = request.get_json() or {}
        cycle = data.get('cycle')
        doc_category = data.get('doc_category')
        directory_path = data.get('directory_path')
        document_patterns = data.get('document_patterns')

        if not all([cycle, doc_category, directory_path]):
            return jsonify({'status': 'error', 'message': '参数不完整'}), 400

        result = user_manager.create_document_directory(
            project_id, cycle, doc_category, directory_path,
            int(current_user.id), document_patterns
        )

        if result['status'] == 'success':
            user_manager.add_operation_log(
                int(current_user.id), current_user.username,
                'create_document_directory', project_id, f'{cycle}/{doc_category}',
                f'directory={directory_path}', ''
            )

        return jsonify(result)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


def delete_document_directory(project_id):
    """删除文档目录映射（仅 PMO/Admin 可用）"""
    try:
        from flask_login import current_user
        from app.models.user import user_manager

        # 权限检查
        if current_user.role not in ('pmo', 'admin'):
            return jsonify({'status': 'error', 'message': '权限不足，仅 PMO/Admin 可操作'}), 403

        data = request.get_json() or {}
        mapping_id = data.get('mapping_id')

        if not mapping_id:
            return jsonify({'status': 'error', 'message': '缺少映射ID'}), 400

        result = user_manager.delete_document_directory(mapping_id)

        if result['status'] == 'success':
            user_manager.add_operation_log(
                int(current_user.id), current_user.username,
                'delete_document_directory', project_id, '',
                f'mapping_id={mapping_id}', ''
            )

        return jsonify(result)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


def preview_excel_file():
    """预览Excel文件内容 - 用户可自定义行列含义"""
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'status': 'error', 'message': '未获取到文件'}), 400
        
        # 保存临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        
        # 读取Excel文件
        wb = openpyxl.load_workbook(tmp_path)
        sheets = wb.sheetnames
        
        # 获取第一个sheet的前20行数据作为预览
        ws = wb[sheets[0]]
        preview_data = []
        max_col = 0
        
        for idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
            preview_data.append({
                'row_num': idx,
                'values': list(row)
            })
            max_col = max(max_col, len([v for v in row if v is not None]))
        
        # 删除临时文件
        Path(tmp_path).unlink(missing_ok=True)
        
        return jsonify({
            'status': 'success',
            'sheets': sheets,
            'preview': preview_data,
            'max_col': max_col
        })
    except Exception as e:
        logger.error(f"预览Excel文件失败: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


def parse_excel_with_mapping():
    """根据用户定义的列映射解析Excel文件"""
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'status': 'error', 'message': '未获取到文件'}), 400
        
        # 获取列映射配置
        mapping = request.form.get('mapping')
        if not mapping:
            return jsonify({'status': 'error', 'message': '未指定列映射'}), 400
        
        try:
            mapping = json.loads(mapping)
        except json.JSONDecodeError:
            return jsonify({'status': 'error', 'message': '列映射格式错误'}), 400
        
        # 保存临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        
        # 读取Excel文件
        wb = openpyxl.load_workbook(tmp_path)
        sheet_name = request.form.get('sheet_name', wb.sheetnames[0])
        ws = wb[sheet_name]
        
        # 根据映射解析数据
        header_row = int(request.form.get('header_row', 1))
        start_row = int(request.form.get('start_row', 2))
        
        parsed_data = {
            'cycles': {},
            'requirements': [],
            'mapping_used': mapping
        }
        
        # 读取数据
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
            row_data = {}
            
            # 根据映射填充数据
            for col_letter, field_name in mapping.items():
                # 将列字母转换为列号 (A=1, B=2...)
                col_num = ord(col_letter.upper()) - ord('A') + 1
                if col_num <= len(row):
                    value = row[col_num - 1]
                    if value is not None:
                        row_data[field_name] = str(value).strip()
            
            # 如果有周期字段，分组数据
            if 'cycle' in row_data:
                cycle = row_data['cycle']
                if cycle not in parsed_data['cycles']:
                    parsed_data['cycles'][cycle] = []
                parsed_data['cycles'][cycle].append(row_data)
            else:
                parsed_data['requirements'].append(row_data)
        
        # 删除临时文件
        Path(tmp_path).unlink(missing_ok=True)
        
        return jsonify({
            'status': 'success',
            'data': parsed_data,
            'rows_processed': row_idx - start_row + 1
        })
    except Exception as e:
        logger.error(f"解析Excel文件失败: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

